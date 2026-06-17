"""
스승의 날 카드뉴스 생성기 - 배송 지연 에디션
단일 반 HTML 생성: python generate.py 1-1
"""

import argparse
import base64
import json
import math
import os
import re
import sys
from pathlib import Path

from google import genai
from google.genai import types as genai_types
import openpyxl
import yaml
from dotenv import load_dotenv
from jinja2 import Environment, FileSystemLoader
from PIL import Image

load_dotenv()

BASE = Path(__file__).parent
TEMPLATES = BASE / "templates"
ANALYSIS = BASE / "analysis"
DATA = BASE / "data"
IMAGES = BASE / "images"
DOCS = BASE / "docs"

for d in [ANALYSIS, DOCS]:
    d.mkdir(exist_ok=True)


# ══════════════════════════════════════════════════
# 설정 로드
# ══════════════════════════════════════════════════

def load_config() -> dict:
    with open(BASE / "config.yaml", encoding="utf-8") as f:
        return yaml.safe_load(f)


# ══════════════════════════════════════════════════
# xlsx 파싱
# ══════════════════════════════════════════════════

def parse_xlsx(class_id: str, acrostic_chars: list) -> dict:
    xlsx_path = DATA / f"{class_id}.xlsx"
    if not xlsx_path.exists():
        print(f"  [경고] {xlsx_path} 없음 - 더미 데이터 사용")
        return _dummy_data(class_id, acrostic_chars)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # 컬럼 인덱스 탐색 (부분 문자열 매칭으로 MS Forms 긴 제목 처리)
    col = {}
    for i, h in enumerate(headers):
        hl = h.lower()
        if any(k in h for k in ("이름", "성명")):
            col.setdefault("name", i)
        elif any(k in hl for k in ("좋은 점", "매력", "좋은점")):
            col.setdefault("charm", i)
        elif any(k in h for k in ("감사", "편지")):
            col.setdefault("letter", i)
        elif any(k in h for k in ("그려", "그리기")):
            col.setdefault("drawing", i)
        elif "ccl" in hl or "라이선스" in h:
            col.setdefault("ccl", i)
        elif acrostic_chars and h.rstrip("!").strip() in acrostic_chars:
            char = h.rstrip("!").strip()
            col[f"acrostic_{char}"] = i

    students = []
    ccl_set = set()
    from collections import Counter
    ccl_counter = Counter()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        name = str(row[col.get("name", 0)] or "").strip()
        if not name:
            continue

        charm = str(row[col.get("charm", -1)] or "") if "charm" in col else ""
        letter = str(row[col.get("letter", -1)] or "") if "letter" in col else ""
        drawing_url = str(row[col.get("drawing", -1)] or "") if "drawing" in col else ""
        ccl = str(row[col.get("ccl", -1)] or "") if "ccl" in col else "CC BY-NC-SA 4.0"

        lines = []
        for char in acrostic_chars:
            key = f"acrostic_{char}"
            val = str(row[col.get(key, -1)] or "") if key in col else ""
            lines.append((char, val))

        if ccl:
            ccl_set.add(ccl)
            ccl_counter[ccl] += 1

        students.append({
            "name": name,
            "charm": charm,
            "letter": letter,
            "drawing_url": drawing_url,
            "ccl": ccl,
            "lines": lines,
        })

    wb.close()
    return {
        "students": students,
        "ccl_licenses": sorted(ccl_set) or ["CC BY-NC-SA 4.0"],
        "ccl_counts": dict(ccl_counter),
    }


def _dummy_data(class_id: str, acrostic_chars: list) -> dict:
    chars = acrostic_chars or ["이", "서", "영"]
    students = []
    names = ["김민준", "이서연", "박지호", "최유나", "정도윤", "강하은", "윤성민", "장윤슬",
             "임수빈", "오준혁", "신아름", "한지원", "류민서", "나동현", "조예린",
             "배현우", "고서윤", "문진호", "권나연", "심재원", "허지수", "남승민", "안소희"]
    for i, name in enumerate(names[:23]):
        lines = []
        for c in chars:
            lines.append((c, f"{c} 자로 시작하는 선생님의 멋진 모습"))
        students.append({
            "name": name,
            "charm": f"항상 밝게 웃어주시는 {name} 눈에 비친 선생님의 매력",
            "letter": f"선생님, 안녕하세요. 저는 {name}입니다. 항상 열심히 가르쳐 주셔서 감사합니다.",
            "drawing_url": "",
            "ccl": "CC BY-NC-SA 4.0",
            "lines": lines,
        })
    return {"students": students, "ccl_licenses": ["CC BY-NC-SA 4.0"]}


# ══════════════════════════════════════════════════
# Claude API 분석
# ══════════════════════════════════════════════════

def run_analysis(class_id: str, students: list, acrostic_chars: list, force: bool = False) -> dict:
    out_path = ANALYSIS / f"{class_id}_analysis.json"

    if out_path.exists() and not force:
        print(f"  [캐시] {out_path} 재사용")
        with open(out_path, encoding="utf-8") as f:
            return json.load(f)

    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key:
        print("  [경고] GEMINI_API_KEY 없음 - 더미 분석 사용")
        return _dummy_analysis(students, acrostic_chars)

    client = genai.Client(api_key=api_key)

    def ask(prompt: str) -> str:
        resp = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=prompt,
        )
        return resp.text.strip()

    all_charms = "\n".join(f"- {s['name']}: {s['charm']}" for s in students if s["charm"])
    all_letters = "\n".join(f"- {s['name']}: {s['letter']}" for s in students if s["letter"])
    all_lines = []
    for s in students:
        for char, line in s["lines"]:
            if line:
                all_lines.append(f"{char}! {line}")

    result = {}

    # 1. 삼행시 연결 스토리
    print("  [API] 삼행시 스토리 생성...")
    try:
        text = ask(f"""다음은 학생들이 쓴 삼행시 줄들입니다. 이를 자연스럽게 연결해 하나의 감동적인 짧은 스토리(3~5문장)로 만들어주세요. 선생님께 드리는 편지처럼 따뜻하게, 배송 지연 유머 컨셉을 살짝 섞어도 좋습니다. JSON {{"story": "..."}}으로만 답해주세요.

삼행시:
{chr(10).join(all_lines[:60])}""")
        data = json.loads(re.search(r'\{.*\}', text, re.DOTALL).group())
        result["story"] = data.get("story", "")
    except Exception as e:
        print(f"    오류: {e}")
        result["story"] = "학생들의 삼행시가 모여 하나의 따뜻한 이야기가 되었습니다. 비록 한 달 늦게 도착했지만, 마음만큼은 스승의 날 그 순간 그대로입니다."

    # 2. 감정 워드클라우드 데이터
    print("  [API] 감정 워드클라우드 분석...")
    try:
        text = ask(f"""다음 학생들의 편지와 매력포인트를 분석해 감정 워드클라우드 데이터를 만들어주세요.
20~30개 단어를, size 1~5(5가 가장 큼), category는 "감사/애정/존경/기쁨/안정감" 중 하나로.
JSON 배열 [{{"word":"","size":1,"category":""}},...] 형식으로만 답해주세요.

편지:
{all_letters[:1500]}

매력포인트:
{all_charms[:800]}""")
        arr = json.loads(re.search(r'\[.*\]', text, re.DOTALL).group())
        result["wordcloud"] = arr
    except Exception as e:
        print(f"    오류: {e}")
        result["wordcloud"] = _dummy_wordcloud()

    # 3. 능력 카테고리 분류
    print("  [API] 능력 카테고리 분류...")
    try:
        text = ask(f"""다음 학생들의 선생님 매력포인트를 4개 카테고리(공감, 관계, 정서, 매력)로 분류하여 각 5~8개 키워드로 추출해주세요.
JSON {{"공감":[],"관계":[],"정서":[],"매력":[]}} 형식으로만 답해주세요.

매력포인트:
{all_charms[:2000]}""")
        data = json.loads(re.search(r'\{.*\}', text, re.DOTALL).group())
        result["abilities"] = {k: data.get(k, []) for k in ["공감", "관계", "정서", "매력"]}
    except Exception as e:
        print(f"    오류: {e}")
        result["abilities"] = {"공감": ["경청해주심", "마음을 알아주심"], "관계": ["친근하심", "소통 잘 하심"], "정서": ["안정감", "따뜻함"], "매력": ["유머", "열정"]}

    # 4. 영향력 분석
    print("  [API] 영향력 분석...")
    try:
        text = ask(f"""다음 학생 편지에서 선생님이 학생들에게 미치는 영향 4가지를 분석해주세요.
각 항목은 title(짧은 제목)과 desc(2~3문장 설명)로 구성.
가장 인상적인 학생 한마디(quote)도 포함.
JSON {{"influences":[{{"title":"","desc":""}},...], "quote":""}} 형식으로만 답해주세요.

편지:
{all_letters[:2000]}""")
        data = json.loads(re.search(r'\{.*\}', text, re.DOTALL).group())
        result["influences"] = data.get("influences", [])
        result["quote"] = data.get("quote", "")
    except Exception as e:
        print(f"    오류: {e}")
        result["influences"] = [
            {"title": "자신감 향상", "desc": "선생님의 격려로 학생들이 스스로를 믿게 되었습니다."},
            {"title": "공부 흥미 증가", "desc": "재미있는 수업 방식 덕분에 배움이 즐거워졌습니다."},
            {"title": "정서적 안정", "desc": "어려울 때마다 따뜻하게 돌봐주셔서 힘이 났습니다."},
            {"title": "관계 능력 성장", "desc": "친구들과 어떻게 지내야 하는지 보여주셨습니다."},
        ]
        result["quote"] = "선생님이 있어서 학교 오는 게 행복해요."

    # 5. 마인드맵 브랜치
    print("  [API] 마인드맵 분석...")
    try:
        text = ask(f"""다음 편지를 5가지 감정 카테고리로 분류하여 각 3~5개 핵심 문구를 추출해주세요.
JSON {{"감사함":[],"사랑함":[],"존경심":[],"기대감":[],"성장의지":[]}} 형식으로만 답해주세요.

편지:
{all_letters[:2000]}""")
        data = json.loads(re.search(r'\{.*\}', text, re.DOTALL).group())
        result["mindmap"] = data
    except Exception as e:
        print(f"    오류: {e}")
        result["mindmap"] = {
            "감사함": ["항상 도와주셔서", "열심히 가르쳐 주셔서", "기다려 주셔서"],
            "사랑함": ["웃음 가득한 교실", "함께여서 행복해요"],
            "존경심": ["최선을 다하시는 모습", "열정적인 수업"],
            "기대감": ["내년에도 함께", "더 많이 배우고 싶어요"],
            "성장의지": ["더 열심히 할게요", "선생님처럼 되고 싶어요"],
        }

    # 6. CCL 성향 분석
    print("  [API] CCL 성향 분석...")
    try:
        ccl_list = [s["ccl"] for s in students if s.get("ccl")]
        from collections import Counter
        ccl_c = Counter(ccl_list)
        ccl_summary = "\n".join(f"- {k}: {v}명" for k, v in ccl_c.most_common())
        text = ask(f"""다음은 중학교 1학년 학생들이 선택한 CCL(Creative Commons License) 라이선스 통계입니다:
{ccl_summary}

이 데이터를 바탕으로 이 학급 학생들의 저작권 성향을 2~3문장으로 분석해주세요.
유머러스하고 재치 있는 톤으로, 마치 MBTI 성격 분석하듯 학생들의 CCL 선택 패턴을 재미있게 해석해주세요.
예: "이 반은 '공유는 하되 돈은 NO' 성향이 강한..." / "저작권계의 MZ세대답게..." 같은 식으로 웃기게 써주세요.
JSON {{"analysis": "..."}} 형식으로만 답해주세요.""")
        data = json.loads(re.search(r'\{.*\}', text, re.DOTALL).group())
        result["ccl_analysis"] = data.get("analysis", "")
    except Exception as e:
        print(f"    오류: {e}")
        result["ccl_analysis"] = "학생들이 직접 CCL 라이선스를 선택하며 자신의 저작물에 대한 권리를 이해하고 실천했습니다."

    # drawing_bullets는 이미지 없으면 기본값
    result["drawing_bullets"] = [
        "대부분의 학생이 선생님을 밝은 표정으로 그렸습니다",
        "선생님의 특징적인 헤어스타일이 많은 그림에서 등장합니다",
        "교실이나 칠판을 배경으로 그린 작품이 많았습니다",
        "다양한 색채로 선생님의 따뜻함을 표현했습니다",
        "학생들과 함께 있는 모습을 그린 작품도 있습니다",
        "하트나 별 등 긍정적인 기호를 함께 그린 학생이 많습니다",
    ]

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"  [저장] {out_path}")
    return result


def _dummy_analysis(students, acrostic_chars):
    return {
        "story": "학생들의 삼행시가 모여 하나의 따뜻한 이야기가 되었습니다.",
        "wordcloud": _dummy_wordcloud(),
        "abilities": {"공감": ["경청해주심"], "관계": ["친근하심"], "정서": ["안정감"], "매력": ["유머"]},
        "influences": [
            {"title": "자신감 향상", "desc": "선생님의 격려로 학생들이 스스로를 믿게 되었습니다."},
            {"title": "공부 흥미 증가", "desc": "재미있는 수업 방식 덕분에 배움이 즐거워졌습니다."},
            {"title": "정서적 안정", "desc": "어려울 때마다 따뜻하게 돌봐주셔서 힘이 났습니다."},
            {"title": "관계 능력 성장", "desc": "친구들과 어떻게 지내야 하는지 보여주셨습니다."},
        ],
        "quote": "선생님이 있어서 학교 오는 게 행복해요.",
        "mindmap": {
            "감사함": ["항상 도와주셔서", "열심히 가르쳐 주셔서"],
            "사랑함": ["웃음 가득한 교실"],
            "존경심": ["최선을 다하시는 모습"],
            "기대감": ["내년에도 함께"],
            "성장의지": ["더 열심히 할게요"],
        },
        "drawing_bullets": [
            "대부분의 학생이 선생님을 밝은 표정으로 그렸습니다",
            "선생님의 특징적인 헤어스타일이 많은 그림에서 등장합니다",
            "교실이나 칠판을 배경으로 그린 작품이 많았습니다",
            "다양한 색채로 선생님의 따뜻함을 표현했습니다",
            "학생들과 함께 있는 모습을 그린 작품도 있습니다",
            "하트나 별 등 긍정적인 기호를 함께 그린 학생이 많습니다",
        ],
    }


def _dummy_wordcloud():
    words = [
        ("감사해요", 5, "감사"), ("사랑해요", 5, "애정"), ("존경해요", 4, "존경"),
        ("행복해요", 4, "기쁨"), ("안심돼요", 4, "안정감"), ("믿음", 3, "안정감"),
        ("따뜻해요", 3, "애정"), ("재미있어요", 3, "기쁨"), ("열정", 3, "존경"),
        ("웃음", 3, "기쁨"), ("격려", 2, "감사"), ("배려", 2, "안정감"),
        ("친절", 2, "애정"), ("성실", 2, "존경"), ("유머", 2, "기쁨"),
        ("지혜", 2, "존경"), ("도움", 2, "감사"), ("친근", 1, "애정"),
        ("용기", 1, "기쁨"), ("진심", 1, "감사"), ("함께", 1, "안정감"),
        ("빛", 1, "존경"), ("희망", 1, "기대감"), ("꿈", 1, "기대감"),
    ]
    return [{"word": w, "size": s, "category": c} for w, s, c in words]


# ══════════════════════════════════════════════════
# SVG 생성
# ══════════════════════════════════════════════════

CATEGORY_COLORS = {
    "감사": "#FF4500",
    "애정": "#FF1493",
    "존경": "#7B1FA2",
    "기쁨": "#FFA500",
    "안정감": "#00BCD4",
    "기대감": "#2E7D32",
}

FONT_SIZES = {1: 22, 2: 34, 3: 48, 4: 64, 5: 82}

# 알록달록 무지개 팔레트
RAINBOW_COLORS = [
    "#FF4500", "#FF1493", "#7B1FA2", "#1565C0", "#00897B",
    "#F57F17", "#00838F", "#AD1457", "#4527A0", "#2E7D32",
    "#E65100", "#6A1B9A", "#0277BD", "#00695C", "#558B2F",
    "#D84315", "#880E4F", "#283593", "#004D40", "#33691E",
]


def make_wordcloud_svg(wordcloud_data: list, width=900, height=400) -> str:
    """wordcloud 라이브러리를 사용한 워드클라우드 — 겹침 없음, 범례 좌측"""
    if not wordcloud_data:
        return f'<svg width="{width}" height="{height}"><text x="50%" y="50%" text-anchor="middle" fill="#aaa">데이터 없음</text></svg>'

    LEGEND_W = 85
    wc_w = width - LEGEND_W
    cy = height // 2

    # ── 카테고리 → 색상 매핑 ─────────────────────────
    word_category = {item["word"]: item.get("category", "감사") for item in wordcloud_data}

    def _color_func(word, font_size, position, orientation, random_state=None, **kwargs):
        cat = word_category.get(word, "감사")
        return CATEGORY_COLORS.get(cat, "#FF4500")

    # ── 단어 빈도: size(1~5) → frequency 비례 매핑 ──
    SIZE_TO_FREQ = {1: 10, 2: 25, 3: 50, 4: 80, 5: 120}
    word_freq = {
        item["word"]: SIZE_TO_FREQ.get(item.get("size", 1), 10)
        for item in wordcloud_data
    }

    # ── 한국어 폰트 경로 ─────────────────────────────
    font_candidates = [
        "C:/Windows/Fonts/malgunbd.ttf",
        "C:/Windows/Fonts/malgun.ttf",
        "C:/Windows/Fonts/Hancom Gothic Bold.ttf",
        "C:/Windows/Fonts/HANDotumB.TTF",
        "C:/Windows/Fonts/HANBatangB.TTF",
    ]
    font_path = next((fp for fp in font_candidates if os.path.exists(fp)), None)

    try:
        from wordcloud import WordCloud

        wc = WordCloud(
            width=wc_w * 2,          # 고해상도로 생성 후 표시 크기는 CSS로 조정
            height=height * 2,
            background_color=None,
            mode="RGBA",
            font_path=font_path,
            color_func=_color_func,
            max_words=60,
            prefer_horizontal=0.65,
            relative_scaling=0.55,
            min_font_size=14,
            max_font_size=110,
            collocations=False,
            margin=8,
        )
        wc.generate_from_frequencies(word_freq)
        pil_img = wc.to_image()

        import io as _io
        buf = _io.BytesIO()
        pil_img.save(buf, format="PNG")
        b64 = base64.b64encode(buf.getvalue()).decode()

    except Exception as e:
        print(f"  [경고] wordcloud 생성 오류: {e} — 대체 텍스트 사용")
        # 가장 큰 단어 3개만 중앙에 표시하는 최소 fallback
        items = sorted(wordcloud_data, key=lambda x: -x.get("size", 1))[:8]
        elems = []
        ys = [height * 0.25, height * 0.5, height * 0.75]
        for k, item in enumerate(items[:3]):
            sz = FONT_SIZES.get(item.get("size", 1), 28)
            col = CATEGORY_COLORS.get(item.get("category", "감사"), "#FF4500")
            elems.append(
                f'<text x="{LEGEND_W + wc_w//2}" y="{ys[k]}" '
                f'text-anchor="middle" font-size="{sz}" font-weight="700" fill="{col}" '
                f'font-family="Noto Sans KR, sans-serif">{item["word"]}</text>'
            )
        used_cats_fb = [c for c in CATEGORY_COLORS if any(d.get("category") == c for d in wordcloud_data)]
        row_h = 22
        ltotal = len(used_cats_fb) * row_h
        ly0 = cy - ltotal // 2
        leg_fb = [f'<rect x="0" y="0" width="{LEGEND_W-6}" height="{height}" fill="rgba(255,255,255,0.08)"/>']
        for i2, cat in enumerate(used_cats_fb):
            col2 = CATEGORY_COLORS[cat]
            leg_fb.append(
                f'<circle cx="12" cy="{ly0+i2*row_h+7}" r="6" fill="{col2}"/>'
                f'<text x="22" y="{ly0+i2*row_h+12}" font-size="12" fill="{col2}" '
                f'font-family="Noto Sans KR, sans-serif" font-weight="700">{cat}</text>'
            )
        return (
            f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" '
            f'viewBox="0 0 {width} {height}">'
            + "".join(leg_fb) + "".join(elems) + "</svg>"
        )

    # ── 범례: 좌측 중앙 수직 배열 ─────────────────────
    used_cats = [cat for cat in CATEGORY_COLORS if any(d.get("category") == cat for d in wordcloud_data)]
    n_cats = len(used_cats)
    row_h = 22
    legend_total_h = n_cats * row_h
    legend_y_start = cy - legend_total_h // 2

    legend_items = [
        f'<rect x="0" y="0" width="{LEGEND_W - 6}" height="{height}" '
        f'fill="rgba(255,255,255,0.08)"/>'
    ]
    for i, cat in enumerate(used_cats):
        col = CATEGORY_COLORS[cat]
        ly = legend_y_start + i * row_h
        legend_items.append(
            f'<circle cx="12" cy="{ly + 7}" r="6" fill="{col}"/>'
            f'<text x="22" y="{ly + 12}" font-size="12" fill="{col}" '
            f'font-family="Noto Sans KR, sans-serif" font-weight="700">{cat}</text>'
        )

    return (
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" '
        f'viewBox="0 0 {width} {height}">'
        f'<image href="data:image/png;base64,{b64}" '
        f'x="{LEGEND_W}" y="0" width="{wc_w}" height="{height}" '
        f'preserveAspectRatio="xMidYMid meet"/>'
        + "".join(legend_items)
        + "</svg>"
    )


def make_mindmap_svg(mindmap: dict, teacher_name: str, width=960, height=520) -> str:
    """matplotlib 마인드맵 — 실제 bbox 측정 후 충돌 감지 루프로 겹침 완전 해결"""
    if not mindmap:
        return (f'<svg width="{width}" height="{height}">'
                f'<text x="50%" y="50%" text-anchor="middle" fill="#aaa">데이터 없음</text></svg>')

    try:
        import textwrap
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        from matplotlib import font_manager
        import io as _io

        branch_colors = ["#FF4500", "#FF1493", "#7B1FA2", "#0288D1", "#2E7D32"]
        branches = list(mindmap.items())
        n = len(branches)

        # ── 한국어 폰트 등록 ─────────────────────────
        font_candidates = [
            "C:/Windows/Fonts/malgunbd.ttf",
            "C:/Windows/Fonts/malgun.ttf",
            "C:/Windows/Fonts/Hancom Gothic Bold.ttf",
            "C:/Windows/Fonts/HANDotumB.TTF",
        ]
        font_path = next((fp for fp in font_candidates if os.path.exists(fp)), None)
        if font_path:
            font_manager.fontManager.addfont(font_path)
            fp_obj = font_manager.FontProperties(fname=font_path)
            plt.rcParams["font.family"] = fp_obj.get_name()
        plt.rcParams["axes.unicode_minus"] = False

        dpi    = 150
        YS     = 0.65   # y 압축 (타원 형태)
        BRANCH_R = 0.38
        LEAF_R   = 0.72
        SECTOR_W = 2 * math.pi / n
        MAX_L    = 3
        WRAP_W   = 12   # 한 줄 최대 글자 수
        BRANCH_FS = 10
        LEAF_FS   = 8.0
        PUSH_STEP = 0.06   # 충돌 시 밀어내는 거리
        MAX_ITER  = 25     # 최대 충돌 해소 반복 횟수

        # 이모지·특수 유니코드 제거 (맑은 고딕 미지원 → □ 깨짐 방지)
        _EMOJI_RE = re.compile(
            "[\U0001F300-\U0001F9FF"   # 이모지 전체 블록
            "\U00002600-\U000027BF"    # 기타 기호
            "\U0001FA00-\U0001FAFF"    # 추가 이모지
            "︀-️"            # Variation Selectors
            "​-‏"            # 제로폭 공백 등
            "]+",
            flags=re.UNICODE,
        )
        def clean_text(text: str) -> str:
            return _EMOJI_RE.sub("", text).strip()

        def wrap_leaf(text: str) -> str:
            cleaned = clean_text(text)
            lines = textwrap.wrap(cleaned, width=WRAP_W)[:2]
            return "\n".join(lines) if lines else cleaned

        # ── 초기 잎 위치 계산 ────────────────────────
        branch_info = []   # (ang, bx, by, color, m_leaves, leaf_init_pos)
        for i, (branch, items) in enumerate(branches):
            ang   = 2 * math.pi * i / n - math.pi / 2
            bx    = BRANCH_R * math.cos(ang)
            by    = BRANCH_R * math.sin(ang) * YS
            color = branch_colors[i % len(branch_colors)]
            m_leaves = items[:MAX_L]
            nl = len(m_leaves)
            half_spread = SECTOR_W * 0.30
            init_pos = []
            for j in range(nl):
                la = ang if nl == 1 else \
                     ang + (j - (nl-1)/2) / (nl-1) * half_spread * 2
                lx = LEAF_R * math.cos(la)
                ly = LEAF_R * math.sin(la) * YS
                init_pos.append((lx, ly, la))
            branch_info.append((ang, bx, by, color, m_leaves, init_pos))

        # ── Figure 생성 (충분히 큰 좌표계) ──────────
        fig, ax = plt.subplots(figsize=(width / dpi, height / dpi), dpi=dpi)
        fig.patch.set_alpha(0)
        ax.set_facecolor("none")
        ax.set_xlim(-1.1, 1.1)
        ax.set_ylim(-0.78, 0.78)
        ax.axis("off")
        fig.subplots_adjust(left=0.01, right=0.99, top=0.99, bottom=0.01)

        # ── Phase 1: 잎 텍스트 먼저 배치 ──────────────
        # (충돌 해소 후 선을 그어야 최종 위치에 연결됨)
        leaf_records = []  # {'text_obj', 'angle', 'bx', 'by', 'color'}
        for (ang, bx, by, color, m_leaves, init_pos) in branch_info:
            for (lx, ly, la), leaf in zip(init_pos, m_leaves):
                wrapped = wrap_leaf(leaf)
                t = ax.text(lx, ly, wrapped,
                            ha="center", va="center",
                            fontsize=LEAF_FS, fontweight="semibold", color=color,
                            clip_on=False, zorder=3, multialignment="center",
                            bbox=dict(boxstyle="round,pad=0.36",
                                      facecolor=color, edgecolor=color,
                                      alpha=0.14, linewidth=1.5))
                leaf_records.append({'text_obj': t, 'angle': la,
                                     'bx': bx, 'by': by, 'color': color})

        # ── Phase 2: 실제 bbox 측정 → 충돌 해소 루프 ─
        fig.canvas.draw()
        renderer  = fig.canvas.get_renderer()
        inv_trans = ax.transData.inverted()

        for _iter in range(MAX_ITER):
            bboxes = [
                r['text_obj'].get_window_extent(renderer).transformed(inv_trans)
                for r in leaf_records
            ]
            moved = False
            for i in range(len(leaf_records)):
                for j in range(i + 1, len(leaf_records)):
                    if bboxes[i].overlaps(bboxes[j]):
                        moved = True
                        # 두 노드 모두 각자의 각도 방향으로 밀어냄
                        for k in (i, j):
                            t   = leaf_records[k]['text_obj']
                            la  = leaf_records[k]['angle']
                            cx, cy = t.get_position()
                            r_cur = math.sqrt(cx**2 + (cy / YS)**2)
                            r_new = r_cur + PUSH_STEP
                            t.set_position((r_new * math.cos(la),
                                            r_new * math.sin(la) * YS))
            if moved:
                fig.canvas.draw()
                renderer = fig.canvas.get_renderer()
            else:
                break

        # ── Phase 3: 최종 위치로 연결선 그리기 (zorder=1) ─
        # 중앙 → 브랜치 선
        for (ang, bx, by, color, m_leaves, _) in branch_info:
            ax.plot([0, bx], [0, by],
                    color=color, lw=2.0, alpha=0.40,
                    linestyle="--", solid_capstyle="round", zorder=1)

        # 브랜치 → 잎 선 (충돌 해소 후 최종 위치 사용)
        for r in leaf_records:
            fx, fy = r['text_obj'].get_position()
            ax.plot([r['bx'], fx], [r['by'], fy],
                    color=r['color'], lw=1.3, alpha=0.40, zorder=1)

        # ── Phase 4: 브랜치 노드 (zorder=4) ─────────
        for (ang, bx, by, color, m_leaves, _), (branch, _) in \
                zip(branch_info, branches):
            ax.text(bx, by, clean_text(branch),
                    ha="center", va="center",
                    fontsize=BRANCH_FS, fontweight="bold", color="white",
                    clip_on=False, zorder=4,
                    bbox=dict(boxstyle="round,pad=0.44",
                              facecolor=color, edgecolor="none", linewidth=0))

        # ── Phase 5: 중앙 원 노드 (zorder=5) ────────
        ax.add_patch(plt.Circle((0, 0), radius=0.105,
                                color="#C2185B", zorder=5,
                                transform=ax.transData))
        ax.text(0,  0.020, f"{teacher_name} 선생님",
                ha="center", va="center",
                fontsize=7.5, fontweight="bold", color="white",
                clip_on=False, zorder=6)
        ax.text(0, -0.030, "마음 지도",
                ha="center", va="center",
                fontsize=6.0, color="white", alpha=0.88,
                clip_on=False, zorder=6)

        # ── 저장: tight bbox → 잘림 없음 ────────────
        buf = _io.BytesIO()
        fig.savefig(buf, format="png", dpi=dpi,
                    bbox_inches="tight", pad_inches=0.10,
                    transparent=True)
        plt.close(fig)
        buf.seek(0)
        b64 = base64.b64encode(buf.read()).decode()

        return (
            f'<svg xmlns="http://www.w3.org/2000/svg" '
            f'width="{width}" height="{height}" viewBox="0 0 {width} {height}">'
            f'<image href="data:image/png;base64,{b64}" '
            f'x="0" y="0" width="{width}" height="{height}" '
            f'preserveAspectRatio="xMidYMid meet"/>'
            f'</svg>'
        )

    except Exception as e:
        print(f"  [경고] matplotlib 마인드맵 오류: {e}")
        import traceback; traceback.print_exc()
        return (f'<svg width="{width}" height="{height}">'
                f'<text x="50%" y="50%" text-anchor="middle" '
                f'font-size="16" fill="#C2185B">마인드맵 생성 오류: {e}</text></svg>')


def make_ccl_chart_svg(ccl_counts: dict, width=800, height=320) -> str:
    """CCL 선택 통계 가로 막대 그래프 SVG"""
    if not ccl_counts:
        return ""
    total = sum(ccl_counts.values())
    items = sorted(ccl_counts.items(), key=lambda x: -x[1])
    bar_colors = ["#FF4500", "#FF1493", "#7B1FA2", "#0288D1", "#2E7D32", "#FFA500"]
    padding_left = 240
    padding_right = 60
    bar_area = width - padding_left - padding_right
    row_h = 42
    chart_h = len(items) * row_h + 20
    actual_h = min(chart_h + 40, height)

    bars = []
    for i, (label, count) in enumerate(items):
        y = 20 + i * row_h
        bar_w = int(bar_area * count / max(ccl_counts.values()))
        color = bar_colors[i % len(bar_colors)]
        pct = int(count / total * 100)
        # 레이블 줄바꿈 처리 (+ 기준)
        parts = label.replace("저작자 표시", "저작자표시").split("+")
        label_short = " + ".join(p.strip()[:6] for p in parts) if len(parts) > 1 else label[:14]
        bars.append(
            f'<text x="{padding_left - 8}" y="{y + 20}" text-anchor="end" '
            f'font-size="13" fill="#333" font-family="Noto Sans KR, sans-serif" font-weight="500">'
            f'{label_short}</text>'
            f'<rect x="{padding_left}" y="{y + 6}" width="{bar_w}" height="24" '
            f'rx="12" fill="{color}" opacity="0.85"/>'
            f'<text x="{padding_left + bar_w + 8}" y="{y + 22}" '
            f'font-size="13" fill="{color}" font-family="Noto Sans KR, sans-serif" font-weight="700">'
            f'{count}명 ({pct}%)</text>'
        )

    return (
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{actual_h}" '
        f'viewBox="0 0 {width} {actual_h}">'
        + "".join(bars)
        + "</svg>"
    )


# ══════════════════════════════════════════════════
# 이미지 처리
# ══════════════════════════════════════════════════

def load_images(class_id: str, students: list, rotation_overrides: dict = None) -> list:
    """학생 그림 이미지 로드.
    rotation_overrides: {"학생이름": 각도} — PIL rotate(각도, expand=True)
      양수 = 반시계(좌), 음수 = 시계(우). 예: {"김예준 10805": -90, "오시훈 10814": 90}
    """
    img_dir = IMAGES / class_id
    if not img_dir.exists():
        return []

    rotation_overrides = rotation_overrides or {}
    result = []
    exts = {".jpg", ".jpeg", ".png", ".gif", ".webp"}
    img_files = {p.stem.lower(): p for p in img_dir.iterdir() if p.suffix.lower() in exts}

    for s in students:
        name = s["name"]
        matched = None
        name_key = name.lower().replace(" ", "").replace("_", "").replace("-", "")
        for stem, path in img_files.items():
            stem_key = stem.lower().replace(" ", "").replace("_", "").replace("-", "")
            if name_key in stem_key:
                matched = path
                break

        if matched:
            try:
                import io
                img = Image.open(matched)
                # EXIF 방향 보정
                try:
                    from PIL import ExifTags
                    exif = img._getexif()
                    if exif:
                        orient_key = next(k for k, v in ExifTags.TAGS.items() if v == "Orientation")
                        orientation = exif.get(orient_key, 1)
                        if orientation == 3:
                            img = img.rotate(180, expand=True)
                        elif orientation == 6:
                            img = img.rotate(270, expand=True)
                        elif orientation == 8:
                            img = img.rotate(90, expand=True)
                except Exception:
                    pass

                # 수동 회전 오버라이드 (config.yaml rotation_overrides 우선)
                if name in rotation_overrides:
                    img = img.rotate(rotation_overrides[name], expand=True)
                # 오버라이드 없을 때만 가로 이미지 자동 90도 회전
                elif img.width > img.height * 1.2:
                    img = img.rotate(90, expand=True)

                img.thumbnail((600, 600))
                buf = io.BytesIO()
                fmt = "JPEG" if matched.suffix.lower() in {".jpg", ".jpeg"} else "PNG"
                img.save(buf, format=fmt, quality=85)
                b64 = base64.b64encode(buf.getvalue()).decode()
                mime = "image/jpeg" if fmt == "JPEG" else "image/png"
                result.append({"name": name, "data": f"data:{mime};base64,{b64}"})
            except Exception as e:
                print(f"    [경고] {matched} 로드 실패: {e}")
                result.append({"name": name, "data": None})
        else:
            result.append({"name": name, "data": None})

    return result


# ══════════════════════════════════════════════════
# HTML 렌더링
# ══════════════════════════════════════════════════

def render_html(class_id: str, cfg: dict, data: dict, analysis: dict) -> str:
    class_cfg = cfg["classes"].get(class_id, {})
    teacher_name = class_cfg.get("teacher") or f"{class_id}반 담임"
    teacher_subject = class_cfg.get("subject", "")
    acrostic_chars = class_cfg.get("acrostic_chars", [])
    students = data["students"]

    # 삼행시 아이템
    acrostic_items = [
        {"name": s["name"], "lines": s["lines"]}
        for s in students if any(line for _, line in s.get("lines", []))
    ]

    # 이미지 분할 (유효 이미지만, 최대 20개, 페이지당 10개)
    rotation_overrides = class_cfg.get("rotation_overrides", {})
    images = load_images(class_id, students, rotation_overrides)
    valid_images = [img for img in images if img.get("data")][:20]
    images_page1 = valid_images[:10]
    images_page2 = valid_images[10:] if len(valid_images) > 10 else []

    # 워드클라우드 / 마인드맵 SVG
    wc_svg = make_wordcloud_svg(analysis.get("wordcloud", []))
    mm_svg = make_mindmap_svg(analysis.get("mindmap", {}), teacher_name)

    # CCL 차트 SVG
    ccl_counts = data.get("ccl_counts", {})
    ccl_chart_svg = make_ccl_chart_svg(ccl_counts)

    # story: 선생님 이름 + 삼행시 기준 글자 하이라이팅
    story_raw = analysis.get("story", "")
    story_highlighted = story_raw
    if teacher_name and story_highlighted:
        story_highlighted = story_highlighted.replace(
            teacher_name,
            f'<span class="teacher-hl">{teacher_name}</span>'
        )
    # 삼행시 기준 글자 각각 하이라이팅 (단어 단위 아닌 글자 단위)
    for char in acrostic_chars:
        if char and story_highlighted:
            story_highlighted = story_highlighted.replace(
                char,
                f'<span class="acrostic-hl">{char}</span>'
            )

    # 폰트 base64
    font_b64 = ""
    font_path = BASE / "assets" / "fonts" / "TDTDGangGulim.woff2"
    if font_path.exists():
        font_b64 = base64.b64encode(font_path.read_bytes()).decode()

    env = Environment(loader=FileSystemLoader(str(TEMPLATES)))
    tpl = env.get_template("card_news.html.jinja2")

    return tpl.render(
        class_id=class_id,
        school=cfg.get("school", "학교"),
        year=cfg.get("year", 2026),
        teacher_name=teacher_name,
        teacher_subject=teacher_subject,
        sent_date=cfg.get("sent_date", "2026-05-15"),
        actual_date=cfg.get("actual_date", "2026-06-15"),
        student_count=len(students),
        acrostic_chars=acrostic_chars,
        acrostic_items=acrostic_items,
        analysis=analysis,
        story_highlighted=story_highlighted,
        wordcloud_svg=wc_svg,
        mindmap_svg=mm_svg,
        images_page1=images_page1,
        images_page2=images_page2,
        ccl_licenses=data.get("ccl_licenses", ["CC BY-NC-SA 4.0"]),
        ccl_counts=ccl_counts,
        ccl_chart_svg=ccl_chart_svg,
        font_base64=font_b64,
        firebase=cfg.get("firebase", {}),
    )


# ══════════════════════════════════════════════════
# 메인
# ══════════════════════════════════════════════════

def generate(class_id: str, force_analysis: bool = False):
    print(f"\n=== {class_id}반 카드뉴스 생성 ===")

    cfg = load_config()
    if class_id not in cfg.get("classes", {}):
        print(f"[오류] config.yaml에 {class_id}가 없습니다.")
        sys.exit(1)

    acrostic_chars = cfg["classes"][class_id].get("acrostic_chars", [])

    print(f"  [1/4] xlsx 파싱...")
    data = parse_xlsx(class_id, acrostic_chars)
    print(f"        학생 수: {len(data['students'])}명")

    print(f"  [2/4] Claude API 분석...")
    analysis = run_analysis(class_id, data["students"], acrostic_chars, force=force_analysis)

    print(f"  [3/4] HTML 렌더링...")
    html = render_html(class_id, cfg, data, analysis)

    out_path = DOCS / f"{class_id}_card_news.html"
    out_path.write_text(html, encoding="utf-8")
    print(f"  [4/4] 저장: {out_path}")
    print(f"  [완료] 파일 크기: {out_path.stat().st_size // 1024} KB")
    return out_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="스승의 날 카드뉴스 생성기")
    parser.add_argument("class_id", help="반 번호 (예: 1-1)")
    parser.add_argument("--force", action="store_true", help="분석 캐시 무시하고 재실행")
    args = parser.parse_args()

    out = generate(args.class_id, force_analysis=args.force)
    print(f"\n브라우저에서 열기: {out}")
